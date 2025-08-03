# excel_utils.py
from openpyxl import load_workbook
from datetime import datetime
from config import EXCEL_FILE, START_SCORE

def load_excel():
    return load_workbook(EXCEL_FILE)

def user_exists(user_id):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == str(user_id):
            wb.close()
            return True
    wb.close()
    return False

def add_participant(user_id, username, team_name):
    wb = load_excel()
    ws = wb["Участники"]
    ws.append(["0", user_id, username, team_name, START_SCORE, "", "", 0])
    wb.save(EXCEL_FILE)
    wb.close()

def get_participants():
    wb = load_excel()
    ws = wb["Участники"]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    wb.close()
    return data

def update_status(user_id, status):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):
            row[0].value = status
            break
    wb.save(EXCEL_FILE)
    wb.close()

def get_status(user_id):
    wb = load_excel()
    ws = wb["Участники"]
    status = None
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):
            status = row[0].value
            break
    wb.close()
    return status

def update_score(user_id, delta):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):
            row[4].value += delta
            break
    wb.save(EXCEL_FILE)
    wb.close()

def set_start_time(user_id):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id) and row[5].value == "":
            row[5].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    wb.save(EXCEL_FILE)
    wb.close()

def set_end_time(user_id):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id) and row[6].value == "":
            row[6].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    wb.save(EXCEL_FILE)
    wb.close()

def add_bonus(user_id, bonus):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):
            row[7].value += bonus
            row[4].value += bonus
            break
    wb.save(EXCEL_FILE)
    wb.close()

def freeze_leader(old_id):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(old_id):
            row[0].value = "заморожен"
            row[3].value = "--"
            row[4].value = 0
            break
    wb.save(EXCEL_FILE)
    wb.close()

def change_leader(old_id, new_id, new_username):
    wb = load_excel()
    ws = wb["Участники"]
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(old_id):
            row[1].value = new_id
            row[2].value = new_username
            break
    wb.save(EXCEL_FILE)
    wb.close()

def get_quest_title():
    wb = load_excel()
    ws = wb["База"]
    title = ws["A1"].value
    wb.close()
    return title

def get_question_data(q_num):
    wb = load_excel()
    ws = wb["База"]
    max_q = 0
    for cell in ws[2][1:]:
        if cell.value:
            max_q += 1
    if q_num > max_q:
        wb.close()
        return None
    col = q_num + 1
    question = ws.cell(2, col).value
    price = ws.cell(3, col).value
    hints = [
        (ws.cell(4, col).value, ws.cell(5, col).value),
        (ws.cell(6, col).value, ws.cell(7, col).value),
        (ws.cell(8, col).value, ws.cell(9, col).value),
    ]
    answers = [a.strip().lower() for a in ws.cell(10, col).value.split(",")]
    wb.close()
    return {"question": question, "price": price, "hints": hints, "answers": answers, "max_q": max_q}
    
def reset_game():
    wb = load_excel()
    ws = wb["Участники"]
    # Удаляем все строки кроме заголовка
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    wb.save(EXCEL_FILE)
    wb.close()

